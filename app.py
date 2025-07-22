import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl

st.set_page_config(page_title="Huella de Carbono - Dashboard", layout="wide")

st.title("📊 Dashboard de Huella de Carbono Organizacional - Sustrend")
st.markdown("Este dashboard muestra **emisiones GEI** e indicadores, basado en los datos de la huella de carbono organizacional del Comité de Desarrollo Productivo Regional Bío-Bío (2024).")

st.sidebar.header("📁 Cargar datos")
uploaded_file = st.sidebar.file_uploader(
    "Sube la planilla Excel de huella de carbono (*.xlsx)",
    type=["xlsx"], help="Cargue archivo Excel con el cálculo oficial."
)

def cargar_datos(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb['Resumen Cálculo']

    emisiones_totales = ws['B30'].value

    indicadores = {}
    row = 46
    while ws[f'B{row}'].value is not None:
        desc = ws[f'A{row}'].value
        valor = ws[f'B{row}'].value
        unidad = ws[f'C{row}'].value
        indicadores[desc] = f"{valor:.2f} {unidad}" if isinstance(valor, (int, float)) else f"{valor} {unidad}"
        row += 1

    xls = pd.ExcelFile(file)
    resumen_df = pd.read_excel(xls, sheet_name="Resumen Cálculo", skiprows=2, usecols="A:C")
    resumen_df.columns = ["Categoria", "Emision_tCO2e", "Porcentaje"]
    resumen_df["Emision_tCO2e"] = pd.to_numeric(resumen_df["Emision_tCO2e"], errors='coerce')

    emisiones_df = pd.read_excel(xls, sheet_name="Resumen Emisiones", header=1)
    emisiones_df.rename(columns={"Emisión CO2e\n[t]": "Emisiones (tCO2e)"}, inplace=True)
    emisiones_df["Emisiones (tCO2e)"] = pd.to_numeric(emisiones_df["Emisiones (tCO2e)"], errors='coerce')

    return emisiones_totales, indicadores, resumen_df, emisiones_df

if uploaded_file:
    try:
        emisiones_totales, indicadores, resumen_df, emisiones_df = cargar_datos(uploaded_file)
    except Exception as e:
        st.error(f"⚠️ Error al leer archivo: {e}")
        st.stop()
else:
    st.warning("⚠️ Por favor, carga un archivo Excel válido.")
    st.stop()

st.header("Resultados del Inventario 2024")
st.metric("Emisiones totales", f"{emisiones_totales:.2f} tCO₂e")

st.subheader("Indicadores de Intensidad Reportados")
cols = st.columns(len(indicadores))
for i, (key, val) in enumerate(indicadores.items()):
    cols[i].metric(key, val)

st.subheader("Distribución de Emisiones por Alcance")
alcance_map = {
    "Alcance 1": "Emisiones directas",
    "Alcance 2": "Emisiones indirectas de energía",
    "Alcance 3": "Otras emisiones indirectas"
}
alcance_colors = {
    "Alcance 1": "#4CAF50",
    "Alcance 2": "#2196F3",
    "Alcance 3": "#FFC107"
}

alcances_df = resumen_df[
    resumen_df["Categoria"].astype(str).str.startswith("Alcance", na=False)
].copy()
alcances_df["Explicacion"] = alcances_df["Categoria"].map(alcance_map)

fig_alcances = px.bar(
    alcances_df, x="Categoria", y="Emision_tCO2e",
    color="Categoria", color_discrete_map=alcance_colors,
    hover_data={"Categoria": True, "Explicacion": True},
    title="Emisiones por Alcance (con explicación)"
)
fig_alcances.update_traces(
    hovertemplate="<b>%{x}</b><br>Emisiones: %{y:.2f} tCO₂e<br>%{customdata[0]}"
)
fig_alcances.update_layout(
    legend_title_text="Alcance",
    margin=dict(t=50, b=80)
)
st.plotly_chart(fig_alcances, use_container_width=True)

st.subheader("Filtrar y Visualizar Emisiones Detalladas")
alcance_opts = sorted(emisiones_df["Alcance"].dropna().unique())
cat_opts = sorted(emisiones_df["Categoría"].dropna().unique())

col1, col2, col3 = st.columns([4, 4, 2])
with col1:
    alcance_sel = st.multiselect("Filtrar por Alcance:", alcance_opts, default=alcance_opts)
with col2:
    cat_sel = st.multiselect("Filtrar por Categoría:", cat_opts, default=cat_opts)
with col3:
    if st.button("🔄 Restaurar filtros"):
        alcance_sel = alcance_opts
        cat_sel = cat_opts

df_filtrado = emisiones_df[
    (emisiones_df["Alcance"].isin(alcance_sel)) &
    (emisiones_df["Categoría"].isin(cat_sel))
]

fig_detalle = px.bar(
    df_filtrado, x="Fuente de Emisión", y="Emisiones (tCO2e)",
    color="Alcance", color_discrete_map=alcance_colors,
    title="Emisiones Detalladas por Fuente",
    hover_data={"Alcance": True, "Fuente de Emisión": True, "Emisiones (tCO2e)": True}
)
fig_detalle.update_layout(
    legend_title_text="Alcance",
    margin=dict(t=50, b=100),
    xaxis_title="Fuente de Emisión",
    yaxis_title="Emisiones (tCO₂e)"
)
st.plotly_chart(fig_detalle, use_container_width=True)

st.dataframe(df_filtrado, use_container_width=True)

csv_data = df_filtrado.to_csv(index=False).encode('utf-8')
st.download_button("💾 Exportar datos filtrados a CSV", csv_data, "emisiones_filtradas.csv", "text/csv")

st.markdown("---")
st.image("sustrend_logo.png", use_container_width=False, width=200)
st.markdown("<center><em>Dashboard elaborado por Sustrend - Datos Programa HuellaChile</em></center>", unsafe_allow_html=True)
